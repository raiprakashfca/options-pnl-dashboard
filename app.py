import streamlit as st
import pandas as pd
import json
import gspread
import re
from datetime import datetime
from io import BytesIO
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Options Trading P&L Dashboard", layout="wide")

scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds_dict = json.loads(st.secrets["GOOGLE_CREDENTIALS_JSON"])
creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
client = gspread.authorize(creds)

SHEET_ID = "1Siith5tw8m-aNOAcwqG1I7L1e_kt8qBX1OOKuAkCpb4"
sheet = client.open_by_key(SHEET_ID).sheet1

def load_trades():
    records = sheet.get_all_records()
    return pd.DataFrame(records)

def append_trades(new_df):
    for col in ["Date", "Expiry"]:
        if col in new_df.columns:
            new_df[col] = new_df[col].astype(str)
    existing = load_trades()
    updated = pd.concat([existing, new_df], ignore_index=True)
    sheet.clear()
    sheet.update([updated.columns.values.tolist()] + updated.values.tolist())

def export_to_excel(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Summary"
    ws.freeze_panes = ws['B2']  # Freeze top row and first column

    col_headers = df.columns.tolist()
    for i, col in enumerate(col_headers, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    current_row = 2
    start_row = current_row
    grouped = df.groupby('Trade Date')
    pnl_col_idx = df.columns.get_loc("P&L") + 1

    for date, group in grouped:
        group = group.sort_values(by='Status', ascending=True)
        pnl_rows = []
        for _, row in group.iterrows():
            row_data = [val if not (col_headers[i] == "P&L" and row['Status'] == 'Open Position') else None for i, val in enumerate(row.tolist())]
            ws.append(row_data)
            if row['Status'] == 'Closed':
                pnl_rows.append(current_row)
            for col_idx, value in enumerate(row.tolist(), 1):
                col_name = col_headers[col_idx - 1]
                cell = ws.cell(row=current_row, column=col_idx)
                cell.alignment = Alignment(horizontal="right" if col_name in ['Buy_Amt', 'Sell_Amt', 'P&L'] else "center")
                if col_name == "Avg_Buy_Price":
                    cell.fill = PatternFill(start_color="BDD7EE", fill_type="solid")
                elif col_name == "Avg_Sell_Price":
                    cell.fill = PatternFill(start_color="F4CCCC", fill_type="solid")
                elif col_name == "P&L" and pd.notna(value):
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                
            current_row += 1

        col_letter = get_column_letter(pnl_col_idx)
        if pnl_rows:
            subtotal_formula = f"=SUM({col_letter}{pnl_rows[0]}:{col_letter}{pnl_rows[-1]})"
        else:
            subtotal_formula = ""
        subtotal_value = group[group['Status'] == 'Closed']['P&L'].sum()
        subtotal_row = [f"Subtotal for {date}"] + [""] * (len(col_headers) - 2) + [subtotal_formula]
        ws.append(subtotal_row)
        row_num = ws.max_row
        for col_idx, val in enumerate(subtotal_row, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = Font(bold=True)
            if col_idx == pnl_col_idx:
                if subtotal_value > 0:
                    cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
                elif subtotal_value < 0:
                    cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        

    col_letter = get_column_letter(pnl_col_idx)
    grand_total_formula = f"=SUM({col_letter}{start_row}:{col_letter}{ws.max_row})"
    grand_total_value = df[df['Status'] == 'Closed']['P&L'].sum()
    grand_row = ["Grand Total"] + [""] * (len(col_headers) - 2) + [grand_total_formula]
    ws.append(grand_row)
    row_num = ws.max_row
    for col_idx, value in enumerate(grand_row, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(bold=True)
        if col_idx == len(grand_row):
            cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid") if grand_total_value > 0 else PatternFill(start_color="FFC7CE", fill_type="solid")

    for col in ws.columns:
        max_len = max(len(str(cell.value) if cell.value else "") for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(output)
    output.seek(0)
    return output

st.title("📊 Options Trading P&L Dashboard")
tab = st.sidebar.radio("Navigation", ["📤 Upload Trades", "📋 Script-Wise Summary"])

if tab == "📤 Upload Trades":
    uploaded_file = st.file_uploader("Upload your trade Excel file (e.g., TRADES01042025.xlsx)", type="xlsx")
    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file)
            df.columns = df.columns.str.strip()
            match = re.search(r"(\d{8})", uploaded_file.name)
            if not match:
                raise ValueError("Filename must contain a valid date in DDMMYYYY format (e.g., TRADES01042025.xlsx)")
            trade_date = datetime.strptime(match.group(1), "%d%m%Y").date()
            df['Trade Date'] = trade_date
            df = df[['Symbol/ScripId', 'Ser/Exp/Group', 'Strike Price', 'Option Type', 'B/S', 'Quantity', 'Price', 'Trade Date']]
            df.columns = ['Symbol', 'Expiry', 'Strike', 'Type', 'Side', 'Quantity', 'Price', 'Date']
            df = df[(df['Quantity'] > 0) & (df['Price'] > 0)]
            df['Value'] = df['Quantity'] * df['Price']
            append_trades(df)
            st.success("✅ Trade file uploaded and saved to Google Sheets.")
        except Exception as e:
            st.error(f"Error reading file: {e}")

elif tab == "📋 Script-Wise Summary":
    st.header("📋 Script-Wise Summary")
    st.cache_data.clear()
    df = load_trades()
        
    required_columns = {'Symbol', 'Expiry', 'Strike', 'Type', 'Side', 'Quantity', 'Price', 'Value', 'Trade Date'}
    if df.empty or not required_columns.issubset(df.columns):
        st.warning("⚠️ No trade data available or expected columns are missing.")
    else:
        try:
            df['OT'] = df['Type'].map({'CE': 'C', 'PE': 'P'})
            df['Leg'] = df['Symbol'].astype(str) + '_' + df['Expiry'].astype(str) + '_' + df['Strike'].astype(str) + '_' + df['OT']
            df = df.rename(columns={'Date': 'Trade Date'})

            status_df = df.groupby(['Symbol', 'Expiry', 'Strike', 'OT'], as_index=False).agg(
                Net_Qty=('Quantity', lambda x: x[df.loc[x.index, 'Side'] == 'S'].sum() - x[df.loc[x.index, 'Side'] == 'B'].sum()),
                PnL=('Value', lambda x: x[df.loc[x.index, 'Side'] == 'S'].sum() - x[df.loc[x.index, 'Side'] == 'B'].sum())
            )
            status_df['Status'] = status_df['Net_Qty'].apply(lambda x: 'Closed' if x == 0 else 'Open Position')

            detailed_df = df.groupby(['Trade Date', 'Symbol', 'Expiry', 'Strike', 'OT'], as_index=False).agg(
                Buy_Qty=('Quantity', lambda x: x[df.loc[x.index, 'Side'] == 'B'].sum()),
                Buy_Amt=('Value', lambda x: x[df.loc[x.index, 'Side'] == 'B'].sum()),
                Sell_Qty=('Quantity', lambda x: x[df.loc[x.index, 'Side'] == 'S'].sum()),
                Sell_Amt=('Value', lambda x: x[df.loc[x.index, 'Side'] == 'S'].sum()),
                Avg_Buy_Price=('Quantity', lambda x: round(df.loc[x.index, 'Value'][df.loc[x.index, 'Side'] == 'B'].sum() / x[df.loc[x.index, 'Side'] == 'B'].sum(), 2) if x[df.loc[x.index, 'Side'] == 'B'].sum() > 0 else None),
                Avg_Sell_Price=('Quantity', lambda x: round(df.loc[x.index, 'Value'][df.loc[x.index, 'Side'] == 'S'].sum() / x[df.loc[x.index, 'Side'] == 'S'].sum(), 2) if x[df.loc[x.index, 'Side'] == 'S'].sum() > 0 else None)
            )

            detailed_df['Net_Qty'] = detailed_df['Sell_Qty'] - detailed_df['Buy_Qty']
            detailed_df['P&L'] = detailed_df['Sell_Amt'] - detailed_df['Buy_Amt']

            merged = pd.merge(detailed_df, status_df[['Symbol', 'Expiry', 'Strike', 'OT', 'Status']], on=['Symbol', 'Expiry', 'Strike', 'OT'], how='left')
            merged = merged.rename(columns={'OT': 'Type'})
            merged = merged.sort_values(by=['Trade Date', 'Symbol', 'Strike'])

            st.dataframe(merged, use_container_width=True)

            totals = merged[merged['Status'] == 'Closed']['P&L'].sum()
            st.markdown(f"### 💰 Total Realised P&L: `{totals:.2f}`")
            excel_file = export_to_excel(merged)
            st.download_button("📥 Download Excel Summary", excel_file, "PnL_Summary.xlsx")

        except Exception as e:
            st.error(f"⚠️ Error processing summary: {e}")
